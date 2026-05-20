import json
import re
import warnings
from pathlib import Path

import openpyxl

# ✅ imports for drawing label extraction
import zipfile
import xml.etree.ElementTree as ET

# =========================
# USER SETTINGS
# =========================
EXCEL_FILE = "Plate Mill layout_V2 (1).xlsx"
SHEET_NAME = "Layout"
OUTPUT_JSON = "full_yard_layout.json"

# ✅ DO NOT change these 4 bays logic
BAYS_IN_ORDER = ["EF", "DE", "CD", "AC"]

# Pixel grid sizing (must match your frontend assumptions)
CELL_W = 60
CELL_H = 60

# Your rule: first visible header column for each bay should map to bay-number 34
START_BAY_NUMBER = 34

# ✅ hard limit bay numbers to 34..67 only (for main 4 bays)
END_BAY_NUMBER = 67

# Blank cell zone rule: "as per the name of the excel sheet"
USE_SHEETNAME_FOR_BLANKS = True
FALLBACK_BLANK_ZONE = "EMPTY"

LETTERS_G_TO_A = ["G", "F", "E", "D", "C", "B", "A"]
LETTERS_F_TO_A = ["F", "E", "D", "C", "B", "A"]

# ✅ EMU to pixel conversion (Excel drawing offsets are in EMU)
EMU_PER_PX = 9525

# =========================
# ✅ OVERLAP / PACKING SETTINGS (CTL FIX)
# =========================
CTL_BLOCK_MARGIN_Y = 120  # space between blocks when auto-shifting
CTL_BLANK_ZONE_TEXT = ""  # ✅ key fix: CTL blank cells should not spam "Layout" text


# ✅ If your frontend already prints the bin id inside each cell,
# adding per-cell CTL zone-name labels creates heavy text overlap.
# Keep this OFF unless you explicitly want zone-name text rendered as separate labels.
CTL_ADD_ZONE_LABELS = False

# =========================
# ✅ YOUR REQUIREMENT FIX (NEW)
# =========================
# CTL should start from arrow point (below main layout), not go outside right.
CTL_START_BELOW_MAIN = True
CTL_BELOW_MAIN_GAP_PX = 140          # gap below main bays (adjust if needed)
CTL_BETWEEN_CTL_BLOCKS_PX = 80       # space between CTLDE and CTLCD blocks (stacked)


# =========================
# HELPERS
# =========================
def build_merged_map(ws):
    """
    Returns a dict: (row, col) -> (min_row, min_col, max_row, max_col)
    for any cell that is inside a merged range.
    """
    merged_map = {}
    for r in ws.merged_cells.ranges:
        for rr in range(r.min_row, r.max_row + 1):
            for cc in range(r.min_col, r.max_col + 1):
                merged_map[(rr, cc)] = (r.min_row, r.min_col, r.max_row, r.max_col)
    return merged_map


def parse_header_range_start(text: str):
    """
    Supports:
      "EF (2-3)" -> returns 2
      "CD (1-2)" -> returns 1
      "EF34"     -> returns 34
      "DE (10-11)" -> returns 10
    """
    if not text:
        return None

    s = str(text).strip()

    m = re.search(r"\((\d+)\s*-\s*(\d+)\)", s)
    if m:
        return int(m.group(1))

    m = re.search(r"^[A-Z]{2}\s*(\d+)", s)
    if m:
        return int(m.group(1))

    return None


def find_header_row_for_bay(ws, bay: str):
    """
    Finds the row that contains the most cells starting with bay prefix (EF/DE/CD/AC).
    Returns: (header_row, header_cols)
    """
    best = None  # (count, row, cols)
    for r in range(1, ws.max_row + 1):
        cols = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().startswith(bay):
                cols.append(c)
        if cols:
            cand = (len(cols), r, cols)
            if best is None or cand > best:
                best = cand

    if not best:
        raise ValueError(f"Could not find header row for bay {bay}")

    _, header_row, header_cols = best
    return header_row, sorted(header_cols)


def find_all_header_rows_for_bay(ws, bay: str):
    """
    Returns ALL candidate header rows for a given bay prefix.
    Each item: (count, row, sorted_cols)
    """
    candidates = []
    for r in range(1, ws.max_row + 1):
        cols = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().startswith(bay):
                cols.append(c)
        if cols:
            candidates.append((len(cols), r, sorted(cols)))

    # higher count first, then earlier row
    candidates.sort(key=lambda x: (-x[0], x[1]))
    return candidates


def _match_sequence(ws, col, start_row, seq):
    for i, expected in enumerate(seq):
        v = ws.cell(start_row + i, col).value
        s = str(v).strip() if v is not None else None
        if s != expected:
            return False
    return True


def find_rowlabel_column_near_header_any(ws, header_row: int, col_left: int, sequences):
    """
    Find row label column near header: for any allowed sequence.
    sequences: list[list[str]] e.g. [G..A] or [F..A]
    Returns: (label_col, letter_to_row, used_sequence)
    """
    best = None  # (distance, col, start_row, seq)

    for c in range(1, col_left + 1):
        for start in range(1, ws.max_row - 10):
            for seq in sequences:
                if start + len(seq) - 1 > ws.max_row:
                    continue
                if _match_sequence(ws, c, start, seq):
                    dist = min(abs(start - header_row), abs((start + len(seq) - 1) - header_row))
                    cand = (dist, c, start, tuple(seq))
                    if best is None or cand < best:
                        best = cand

    if not best:
        raise ValueError(
            f"Could not infer row label column near header row {header_row}. "
            f"Ensure row labels like {sequences} exist in ONE column."
        )

    _, label_col, start_row, seq = best
    seq = list(seq)
    letter_to_row = {seq[i]: start_row + i for i in range(len(seq))}
    return label_col, letter_to_row, seq


def build_col_to_baynum(ws, bay: str, header_row: int, header_cols):
    """
    MAIN 4 BAYS ONLY:
    Maps each header column to bay number like 34,35,... based on your rule.
    """
    starts = []
    for c in header_cols:
        v = ws.cell(header_row, c).value
        start_n = parse_header_range_start(v)
        if start_n is not None:
            starts.append((c, start_n))

    if not starts:
        raise ValueError(f"No numeric header ranges found for bay {bay} on row {header_row}")

    starts.sort(key=lambda x: x[0])
    first_start = starts[0][1]
    offset = START_BAY_NUMBER - first_start

    col_to_num = {c: (start_n + offset) for c, start_n in starts}
    return col_to_num


def build_col_to_baynum_ctl(ws, header_row: int, header_cols):
    """
    CTL SHOP:
    Map header columns using the ACTUAL numbers in headers like:
      DE (2-3) -> 2
      CD (7-8) -> 7
    No START_BAY_NUMBER shifting, no 34..67 filtering.
    """
    col_to_num = {}
    for c in header_cols:
        v = ws.cell(header_row, c).value
        start_n = parse_header_range_start(v)
        if start_n is not None:
            col_to_num[c] = start_n

    if not col_to_num:
        raise ValueError(f"CTL header row {header_row} has no parseable numeric ranges.")

    return col_to_num


def first_text_in_rect(ws, r1, c1, r2, c2):
    """
    Find the first non-empty text within the rectangle.
    """
    for rr in range(r1, r2 + 1):
        for cc in range(c1, c2 + 1):
            v = ws.cell(rr, cc).value
            if v is None:
                continue
            s = str(v).strip()
            if s:
                return s
    return None


def find_first_cell_exact(ws, text_upper: str):
    """
    Returns (row, col) for first cell whose stripped upper() equals text_upper.
    """
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().upper() == text_upper:
                return r, c
    return None


def row_has_numeric_ranges(ws, row: int, cols) -> bool:
    """
    True if at least one of the given cells looks like 'DE (2-3)' / 'CD (7-8)'.
    """
    for c in cols:
        v = ws.cell(row, c).value
        if isinstance(v, str) and re.search(r"\(\s*\d+\s*-\s*\d+\s*\)", v.strip()):
            return True
    return False


def pick_ctl_header_row(ws, candidates, anchor_row: int, anchor_col: int | None, expect_prefix: str, search_down_rows: int = 12):
    """
    Picks the correct CTL header row near an anchor.
    Strategy:
      - Look in window [anchor_row .. anchor_row+search_down_rows]
      - Among candidates in this window, choose the first row that:
           (a) has numeric ranges
           (b) has cells starting with expect_prefix (DE/CD)
    """
    window_min = anchor_row
    window_max = min(ws.max_row, anchor_row + search_down_rows)

    valid = []
    for count, row, cols in candidates:
        if not (window_min <= row <= window_max):
            continue
        if not row_has_numeric_ranges(ws, row, cols):
            continue

        ok = False
        for c in cols:
            v = ws.cell(row, c).value
            if isinstance(v, str) and v.strip().startswith(expect_prefix):
                ok = True
                break
        if not ok:
            continue

        valid.append((row, -count, cols))

    if not valid:
        return None

    valid.sort(key=lambda x: (x[0], x[1]))
    row, _, cols = valid[0]
    return row, cols


# =========================
# ✅ STYLE EXTRACTION
# =========================
def _argb_to_hex(argb: str | None):
    if not argb:
        return None
    s = str(argb).strip()
    if not s:
        return None
    if len(s) == 8:
        s = s[2:]
    if len(s) != 6:
        return None
    return f"#{s.upper()}"


def _safe_color(color_obj):
    if color_obj is None:
        return None
    rgb = getattr(color_obj, "rgb", None)
    if rgb:
        return _argb_to_hex(rgb)
    return None


def extract_cell_style(cell):
    fill_hex = None
    try:
        fill = cell.fill
        if fill and getattr(fill, "patternType", None) == "solid":
            fill_hex = _safe_color(getattr(fill, "fgColor", None))
    except Exception:
        fill_hex = None

    font_color_hex = None
    font_bold = None
    font_italic = None
    font_name = None
    font_size = None
    try:
        font = cell.font
        if font:
            font_color_hex = _safe_color(getattr(font, "color", None))
            font_bold = bool(getattr(font, "b", False))
            font_italic = bool(getattr(font, "i", False))
            font_name = getattr(font, "name", None)
            font_size = getattr(font, "sz", None)
    except Exception:
        pass

    align = {}
    try:
        al = cell.alignment
        if al:
            align = {
                "horizontal": getattr(al, "horizontal", None),
                "vertical": getattr(al, "vertical", None),
                "wrap_text": bool(getattr(al, "wrap_text", False)),
                "text_rotation": getattr(al, "text_rotation", None),
            }
    except Exception:
        align = {}

    def side_to_dict(side):
        if side is None:
            return {"style": None, "color": None}
        return {"style": getattr(side, "style", None), "color": _safe_color(getattr(side, "color", None))}

    border = {}
    try:
        b = cell.border
        if b:
            border = {
                "left": side_to_dict(getattr(b, "left", None)),
                "right": side_to_dict(getattr(b, "right", None)),
                "top": side_to_dict(getattr(b, "top", None)),
                "bottom": side_to_dict(getattr(b, "bottom", None)),
            }
    except Exception:
        border = {}

    return {
        "fill": fill_hex,
        "font_color": font_color_hex,
        "font_bold": font_bold,
        "font_italic": font_italic,
        "font_name": font_name,
        "font_size": font_size,
        "alignment": align,
        "border": border,
    }


# =========================
# ✅ EXTRACT SHAPE/TEXTBOX LABELS
# =========================
def _sheet_rid_and_target_drawing(excel_path: Path, sheet_name: str):
    with zipfile.ZipFile(excel_path, "r") as z:
        wb_xml = ET.fromstring(z.read("xl/workbook.xml"))
        ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        sheets = wb_xml.find("m:sheets", ns)
        if sheets is None:
            return None

        target_rid = None
        for sh in sheets.findall("m:sheet", ns):
            if sh.attrib.get("name") == sheet_name:
                target_rid = sh.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
                break

        if not target_rid:
            return None

        rels_xml = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}

        sheet_path = None
        for rel in rels_xml.findall("r:Relationship", rns):
            if rel.attrib.get("Id") == target_rid:
                sheet_path = "xl/" + rel.attrib.get("Target").lstrip("/")
                break

        if not sheet_path or sheet_path not in z.namelist():
            return None

        sheet_rels_path = str(Path(sheet_path).parent / "_rels" / (Path(sheet_path).name + ".rels"))
        if sheet_rels_path not in z.namelist():
            return None

        sheet_rels_xml = ET.fromstring(z.read(sheet_rels_path))

        drawing_rid = None
        for rel in sheet_rels_xml.findall("r:Relationship", rns):
            if rel.attrib.get("Type", "").endswith("/drawing"):
                drawing_rid = rel.attrib.get("Id")
                break

        if not drawing_rid:
            return None

        sheet_xml = ET.fromstring(z.read(sheet_path))
        drawing_el = sheet_xml.find(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}drawing")
        if drawing_el is None:
            return None

        sheet_drawing_rid = drawing_el.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        if sheet_drawing_rid != drawing_rid:
            drawing_rid = sheet_drawing_rid

        for rel in sheet_rels_xml.findall("r:Relationship", rns):
            if rel.attrib.get("Id") == drawing_rid:
                target = rel.attrib.get("Target")
                if not target:
                    return None
                drawing_path = "xl/" + target.lstrip("/")
                if drawing_path in z.namelist():
                    return drawing_path

        return None


def extract_drawing_text_labels(excel_path: Path, sheet_name: str, x_from_col, y_from_row):
    drawing_path = _sheet_rid_and_target_drawing(excel_path, sheet_name)
    if not drawing_path:
        return []

    labels = []

    with zipfile.ZipFile(excel_path, "r") as z:
        root = ET.fromstring(z.read(drawing_path))
        ns = {
            "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
            "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
        }

        anchors = (
            root.findall("xdr:twoCellAnchor", ns)
            + root.findall("xdr:oneCellAnchor", ns)
            + root.findall("xdr:absoluteAnchor", ns)
        )

        for anc in anchors:
            fr = anc.find("xdr:from", ns)
            if fr is None:
                continue

            c_el = fr.find("xdr:col", ns)
            r_el = fr.find("xdr:row", ns)
            if c_el is None or r_el is None:
                continue

            col0 = int(c_el.text or "0") + 1
            row0 = int(r_el.text or "0") + 1

            col_off = fr.find("xdr:colOff", ns)
            row_off = fr.find("xdr:rowOff", ns)
            off_x_px = (int(col_off.text) / EMU_PER_PX) if (col_off is not None and col_off.text) else 0.0
            off_y_px = (int(row_off.text) / EMU_PER_PX) if (row_off is not None and row_off.text) else 0.0

            texts = [t.text for t in anc.findall(".//a:t", ns) if t.text]
            if not texts:
                continue

            text = " ".join([s.strip() for s in texts if s.strip()]).strip()
            if not text:
                continue

            x = x_from_col(col0) + off_x_px
            y = y_from_row(row0) + off_y_px

            labels.append({"text": text, "x_pos": int(x), "y_pos": int(y), "width": None})

    return labels


# =========================
# ✅ RECT / OVERLAP UTILITIES
# =========================
def rect_from_zone(z):
    x1 = int(z["x_pos"])
    y1 = int(z["y_pos"])
    x2 = x1 + int(z["width"])
    y2 = y1 + int(z["height"])
    return (x1, y1, x2, y2)


def rects_intersect(a, b) -> bool:
    ax1, ay1, ax2, ay2 = a
    bx1, by1, bx2, by2 = b
    return (ax1 < bx2) and (ax2 > bx1) and (ay1 < by2) and (ay2 > by1)


def bbox_of_zones(zones):
    if not zones:
        return None
    xs1, ys1, xs2, ys2 = [], [], [], []
    for z in zones:
        x1, y1, x2, y2 = rect_from_zone(z)
        xs1.append(x1); ys1.append(y1); xs2.append(x2); ys2.append(y2)
    return (min(xs1), min(ys1), max(xs2), max(ys2))


def shift_block(block_zones, block_labels, dx=0, dy=0):
    if dx == 0 and dy == 0:
        return
    for z in block_zones:
        z["x_pos"] = int(z["x_pos"]) + dx
        z["y_pos"] = int(z["y_pos"]) + dy
    for l in block_labels:
        l["x_pos"] = int(l["x_pos"]) + dx
        l["y_pos"] = int(l["y_pos"]) + dy


def pack_block_down_if_overlapping(block_zones, block_labels, existing_rects, margin_y=CTL_BLOCK_MARGIN_Y):
    """
    ✅ CTL Overlap Fix:
    If CTL block bbox intersects any already-emitted zone rect,
    shift the entire CTL block DOWN until it no longer intersects.
    """
    if not block_zones:
        return

    # Loop safety
    for _ in range(50):
        bb = bbox_of_zones(block_zones)
        if bb is None:
            return

        overlaps = [r for r in existing_rects if rects_intersect(bb, r)]
        if not overlaps:
            return

        # push block below the lowest overlapping rect
        max_y2 = max(r[3] for r in overlaps)
        needed_dy = (max_y2 + margin_y) - bb[1]
        if needed_dy <= 0:
            needed_dy = margin_y

        shift_block(block_zones, block_labels, dy=needed_dy)

    # If still overlapping after many tries, we still output it (but this should not happen)
    return


# =========================
# MAIN BUILD
# =========================
def main():
    warnings.filterwarnings(
        "ignore",
        message="DrawingML support is incomplete*",
        category=UserWarning,
    )

    excel_path = Path(EXCEL_FILE)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {EXCEL_FILE}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")

    ws = wb[SHEET_NAME]
    merged_map = build_merged_map(ws)
    merged_ranges = [str(rng) for rng in ws.merged_cells.ranges]

    # --- Find header blocks for each of the main 4 bays (UNCHANGED)
    bay_meta = {}
    for bay in BAYS_IN_ORDER:
        header_row, header_cols = find_header_row_for_bay(ws, bay)
        col_left = min(header_cols)

        # ✅ main bays use G..A (keep original)
        _, letter_to_row, _used_seq = find_rowlabel_column_near_header_any(
            ws, header_row, col_left, sequences=[LETTERS_G_TO_A]
        )

        col_to_num_raw = build_col_to_baynum(ws, bay, header_row, header_cols)

        # ✅ keep ONLY bay numbers 34..67 (UNCHANGED)
        col_to_num = {c: n for c, n in col_to_num_raw.items() if START_BAY_NUMBER <= n <= END_BAY_NUMBER}

        bay_meta[bay] = {
            "header_row": header_row,
            "header_cols": header_cols,
            "col_left": col_left,
            "letter_to_row": letter_to_row,
            "col_to_num": col_to_num,
        }

    # --- Anchor EF's G cell position to pixel (100,40) like your earlier JSON (UNCHANGED)
    ef = bay_meta["EF"]
    if not ef["col_to_num"]:
        raise ValueError(
            "After filtering to 34..67, EF has no usable columns. "
            "Check your Excel header ranges and START/END bay numbers."
        )

    EF_MIN_COL = min(ef["header_cols"])
    EF_G_ROW = ef["letter_to_row"]["G"]

    def x_from_col(col_idx: int) -> int:
        return 100 + (col_idx - EF_MIN_COL) * CELL_W

    def y_from_row(row_idx: int) -> int:
        return 40 + (row_idx - EF_G_ROW) * CELL_H

    blank_zone_main = ws.title if USE_SHEETNAME_FOR_BLANKS else FALLBACK_BLANK_ZONE

    zones = []
    labels = []
    used_bins = set()

    # ✅ Prevent duplicate row-letter labels being added multiple times
    row_letter_positions_emitted = set()  # (text, x_pos, y_pos)

    # ✅ for overlap checks (REAL intersection, not "same rect")
    existing_rects = []

    # =========================
    # MAIN 4 BAYS BUILD (UNCHANGED)
    # =========================
    for bay in BAYS_IN_ORDER:
        meta = bay_meta[bay]
        letter_to_row = meta["letter_to_row"]
        col_to_num = meta["col_to_num"]

        header_cols = sorted(col_to_num.keys())
        if not header_cols:
            continue

        min_valid_col = min(header_cols)
        max_valid_col = max(header_cols)

        # Labels: column labels
        g_row = letter_to_row["G"]
        header_y = y_from_row(g_row) - 30
        for c in header_cols:
            num = col_to_num[c]
            labels.append({"text": f"{bay}{num}", "x_pos": x_from_col(c), "y_pos": header_y, "width": CELL_W})


        # Labels: row letters (emit once per unique position)
        for letter, r in letter_to_row.items():
            key = (letter, 20, y_from_row(r))
            if key in row_letter_positions_emitted:
                continue
            row_letter_positions_emitted.add(key)
            labels.append({"text": letter, "x_pos": 20, "y_pos": y_from_row(r), "width": None})


        visited_top_left = set()
        row_by_letter = {r: l for l, r in letter_to_row.items()}
        grid_rows = sorted(letter_to_row.values())

        for r in grid_rows:
            letter = row_by_letter.get(r)
            if not letter:
                continue

            for c in header_cols:
                tl = merged_map.get((r, c), (r, c, r, c))
                tl_row, tl_col, br_row, br_col = tl

                if (r, c) != (tl_row, tl_col):
                    continue
                if (tl_row, tl_col) in visited_top_left:
                    continue
                visited_top_left.add((tl_row, tl_col))

                if tl_col < min_valid_col:
                    continue
                br_col_clamped = min(br_col, max_valid_col)
                if br_col_clamped < tl_col:
                    continue

                found_text = first_text_in_rect(ws, tl_row, tl_col, br_row, br_col_clamped)
                zone_name = found_text if found_text else blank_zone_main

                bay_num = col_to_num[c]
                bin_id = f"{bay}{bay_num}{letter}"

                if bin_id in used_bins:
                    continue
                used_bins.add(bin_id)

                zone_x = x_from_col(tl_col)
                zone_y = y_from_row(tl_row)
                zone_w = (br_col_clamped - tl_col + 1) * CELL_W
                zone_h = (br_row - tl_row + 1) * CELL_H

                style = extract_cell_style(ws.cell(tl_row, tl_col))

                z = {
                    "bin": bin_id,
                    "zone": zone_name,
                    "row": letter,
                    "bay": bay,
                    "x_pos": zone_x,
                    "y_pos": zone_y,
                    "width": zone_w,
                    "height": zone_h,
                    "style": style,
                    "z_index": 20,
                    "merge": {
                        "min_row": tl_row,
                        "min_col": tl_col,
                        "max_row": br_row,
                        "max_col": br_col_clamped,
                    },
                }

                zones.append(z)
                existing_rects.append(rect_from_zone(z))

    # ✅ NEW: compute main layout bbox → CTL anchor (arrow position)
    main_bbox = bbox_of_zones(zones)
    if main_bbox:
        CTL_ANCHOR_X = main_bbox[0]
        CTL_NEXT_Y = main_bbox[3] + CTL_BELOW_MAIN_GAP_PX
    else:
        CTL_ANCHOR_X = 100
        CTL_NEXT_Y = 40 + CTL_BELOW_MAIN_GAP_PX

    # =========================
    # ✅ CTL SHOP BUILD (NO OVERLAP + NO BLANK TEXT SPAM)
    # =========================
    ctl_cell = find_first_cell_exact(ws, "CTL")
    cd_bay_cell = None

    if ctl_cell:
        ctl_r, ctl_c = ctl_cell
        # Find "CD BAY" below CTL to avoid main CD BAY
        for r in range(ctl_r, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip().upper() == "CD BAY":
                    cd_bay_cell = (r, c)
                    break
            if cd_bay_cell:
                break

        de_blocks = find_all_header_rows_for_bay(ws, "DE")
        cd_blocks = find_all_header_rows_for_bay(ws, "CD")

        ctl_de_header = pick_ctl_header_row(ws, de_blocks, ctl_r, ctl_c, expect_prefix="DE", search_down_rows=10)
        if cd_bay_cell:
            cd_r, cd_c = cd_bay_cell
            ctl_cd_header = pick_ctl_header_row(ws, cd_blocks, cd_r, cd_c, expect_prefix="CD", search_down_rows=10)
        else:
            ctl_cd_header = None
    else:
        ctl_de_header = None
        ctl_cd_header = None

    def build_ctl_bay(bay_name: str, header_row: int, header_cols):
        """
        Build CTL bay into TEMP lists first, then auto-pack down if it overlaps
        anything already placed. This guarantees NO CTL overlap in final JSON.

        ✅ NEW: shift CTL below main (arrow position) and stack CTL blocks.
        """
        nonlocal CTL_NEXT_Y

        temp_zones = []
        temp_labels = []
        row_letter_labels = []  # add after shift so duplicate-check uses final y

        col_left = min(header_cols)

        # CTL grid uses F..A (but accept G..A if present)
        _, letter_to_row, used_seq = find_rowlabel_column_near_header_any(
            ws, header_row, col_left, sequences=[LETTERS_F_TO_A, LETTERS_G_TO_A]
        )

        col_to_num = build_col_to_baynum_ctl(ws, header_row, header_cols)
        header_cols_sorted = sorted(col_to_num.keys(), key=lambda c: col_to_num[c])

        # Column labels (CTL-prefixed)
        top_letter = used_seq[0]
        top_row = letter_to_row[top_letter]
        header_y = y_from_row(top_row) - 30
        for c in header_cols_sorted:
            num = col_to_num[c]
            temp_labels.append({"text": f"CTL{bay_name}{num}", "x_pos": x_from_col(c), "y_pos": header_y, "width": CELL_W})


        # Row labels (prepare; add after shift)
        for letter, r in letter_to_row.items():
            row_letter_labels.append({"text": letter, "x_pos": 20, "y_pos": y_from_row(r), "width": None})


        visited_top_left = set()
        row_by_letter = {r: l for l, r in letter_to_row.items()}
        grid_rows = sorted(letter_to_row.values())

        min_valid_col = min(header_cols_sorted)
        max_valid_col = max(header_cols_sorted)

        for r in grid_rows:
            letter = row_by_letter.get(r)
            if not letter:
                continue

            for c in header_cols_sorted:
                tl = merged_map.get((r, c), (r, c, r, c))
                tl_row, tl_col, br_row, br_col = tl

                if (r, c) != (tl_row, tl_col):
                    continue
                if (tl_row, tl_col) in visited_top_left:
                    continue
                visited_top_left.add((tl_row, tl_col))

                if tl_col < min_valid_col:
                    continue
                br_col_clamped = min(br_col, max_valid_col)
                if br_col_clamped < tl_col:
                    continue

                # ✅ CTL FIX: don't use sheet-name for blanks (this creates heavy "overlap" visually)
                found_text = first_text_in_rect(ws, tl_row, tl_col, br_row, br_col_clamped)
                zone_name = found_text if found_text else CTL_BLANK_ZONE_TEXT

                bay_num = col_to_num[c]
                bin_id = f"CTL{bay_name}{bay_num}{letter}"
                if bin_id in used_bins:
                    continue
                used_bins.add(bin_id)

                zone_x = x_from_col(tl_col)
                zone_y = y_from_row(tl_row)
                zone_w = (br_col_clamped - tl_col + 1) * CELL_W
                zone_h = (br_row - tl_row + 1) * CELL_H

                style = extract_cell_style(ws.cell(tl_row, tl_col))

                z = {
                    "bin": bin_id,
                    "zone": zone_name,
                    "row": letter,
                    "bay": "CTL",
                    "x_pos": zone_x,
                    "y_pos": zone_y,
                    "width": zone_w,
                    "height": zone_h,
                    "style": style,
                    "z_index": 5,  # behind main bays
                    "merge": {
                        "min_row": tl_row,
                        "min_col": tl_col,
                        "max_row": br_row,
                        "max_col": br_col_clamped,
                    },
                }
                temp_zones.append(z)
                # ✅ Avoid per-cell CTL zone-name labels by default (they overlap with bin text in UI).
                # If you later want these, set CTL_ADD_ZONE_LABELS = True.
                if CTL_ADD_ZONE_LABELS and found_text:
                    # Center-ish placement: use top-left; your UI can center using width.
                    temp_labels.append({"text": zone_name, "x_pos": zone_x, "y_pos": zone_y, "width": zone_w})

        # ✅ NEW: SHIFT THIS CTL BLOCK BELOW MAIN + ALIGN LEFT (arrow point)
        if CTL_START_BELOW_MAIN and temp_zones:
            bb = bbox_of_zones(temp_zones)
            if bb:
                dx = CTL_ANCHOR_X - bb[0]
                dy = CTL_NEXT_Y - bb[1]
                shift_block(temp_zones, temp_labels, dx=dx, dy=dy)

                # shift row-letter labels ONLY in Y (keep x=20 fixed)
                for rl in row_letter_labels:
                    rl["y_pos"] = int(rl["y_pos"]) + dy

        # ✅ add row letters after shift (so duplicates check is correct)
        for rl in row_letter_labels:
            key = (rl["text"], rl["x_pos"], rl["y_pos"])
            if key in row_letter_positions_emitted:
                continue
            row_letter_positions_emitted.add(key)
            temp_labels.append(rl)

        # ✅ KEY FIX: if CTL block overlaps anything already placed, push it DOWN
        pack_block_down_if_overlapping(temp_zones, temp_labels, existing_rects, margin_y=CTL_BLOCK_MARGIN_Y)

        # ✅ update CTL_NEXT_Y so next CTL block stacks below
        bb2 = bbox_of_zones(temp_zones)
        if bb2:
            CTL_NEXT_Y = bb2[3] + CTL_BETWEEN_CTL_BLOCKS_PX

        # Commit temp -> final
        zones.extend(temp_zones)
        labels.extend(temp_labels)
        for z in temp_zones:
            existing_rects.append(rect_from_zone(z))

    ctl_build_list = []

    if ctl_de_header:
        de_row, de_cols = ctl_de_header
        ctl_build_list.append(("DE", de_row, de_cols))

    if ctl_cd_header:
        cd_row, cd_cols = ctl_cd_header
        ctl_build_list.append(("CD", cd_row, cd_cols))

    # ✅ Build CTL bays in visual order (top to bottom in the sheet)
    ctl_build_list.sort(key=lambda x: x[1])

    prev_row = None
    for bay_name, row, cols in ctl_build_list:
        # if somehow same header row appears twice, skip duplicates
        if prev_row is not None and row == prev_row:
            continue
        build_ctl_bay(bay_name, row, cols)
        prev_row = row

    # Add drawing/textbox labels (if any)
    drawing_labels = extract_drawing_text_labels(excel_path, SHEET_NAME, x_from_col, y_from_row)
    labels.extend(drawing_labels)

    out = {
        "merged_ranges": merged_ranges,
        "zones": zones,
        "labels": labels,
    }

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=2, ensure_ascii=False)

    print(f"✅ Wrote {OUTPUT_JSON}")
    print(f"   zones:  {len(zones)}")
    print(f"   labels: {len(labels)}")
    print(f"   drawing labels added: {len(drawing_labels)}")
    print(f"   merged ranges exported: {len(merged_ranges)}")
    print("   Bays found:", ", ".join(BAYS_IN_ORDER))

    if ctl_cell:
        print(f"   CTL found at row {ctl_cell[0]}, col {ctl_cell[1]}")
    else:
        print("   CTL: NOT FOUND")

    if cd_bay_cell:
        print(f"   CTL CD BAY marker found at row {cd_bay_cell[0]}, col {cd_bay_cell[1]}")
    else:
        print("   CTL CD BAY marker: NOT FOUND")


if __name__ == "__main__":
    main()
