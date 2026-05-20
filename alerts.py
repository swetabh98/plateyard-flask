from __future__ import annotations

import os
import json
import time
import argparse
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import List, Dict, Any, Optional, Tuple

from sqlalchemy import create_engine, text

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

try:
    from twilio.rest import Client as TwilioClient
except Exception:
    TwilioClient = None


# =========================
# CONFIG (hardcoded for now)
# =========================

ALERT_TITLE = "Plate Mill Yard Aging Stock"

# ✅ Your requested recipients (for now)
DEFAULT_WHATSAPP_NUMBERS = ["9109951823"]  # India number without +91

# ✅ TO recipients (as requested)
DEFAULT_EMAIL_RECIPIENTS = [
    "rupak.ranjan@jindalsteel.in",
    "pmlog.shiftincharge.rgh@jindalsteel.in",
    "hemant.patel1@jindalsteel.in",
    "kshitibhushan.gupta@jindalsteel.in",
    "sunilpatel.ex@gmail.com",
    "prashant.chaudhari@jindalsteel.in",
]

# ✅ CC recipients (as requested)
DEFAULT_EMAIL_CC = [
    "lalit.goyal@jindalsteel.in",
    "swetabh.sinha@jindalsteel.in",
    "aritra.de@jindalsteel.in",
    "saurabh.agrawal@jindalsteel.in",
    "man.singh@jindalsteel.in",
    "sandeep.tyagi@jindalsteel.in",
    "vikram.singh1@jindalsteel.in",
    "sushil.mohanty@jindalsteel.in",
    "abhimanyu.singh@jindalsteel.in",
    "devendra.kalambe@jindalsteel.in",
]

# Thresholds (days)
FG_YELLOW_DAYS = 15
FG_RED_DAYS = 20

WIP_YELLOW_DAYS = 3
WIP_RED_DAYS = 7

# =========================
# SMTP (use your provided working style)
# =========================
SENDER_NAME = "Yard Management Software"
SENDER_EMAIL = "noreply.digital@jindalsteel.com"
SMTP_SERVER = "172.17.1.17"
SMTP_PORT = 25

# Simple Jindal-styled colors (email HTML)
BRAND_ORANGE = os.environ.get("JINDAL_BRAND_ORANGE", "#F58220")
BRAND_GREEN  = os.environ.get("JINDAL_BRAND_GREEN",  "#3AAA35")
BRAND_BLUE   = os.environ.get("JINDAL_BRAND_BLUE",   "#0067B2")
BRAND_BG     = os.environ.get("JINDAL_BRAND_MUTE",   "#F4F6F8")
BRAND_TEXT   = os.environ.get("JINDAL_BRAND_TEXT",   "#333333")
BRAND_FONT   = os.environ.get("JINDAL_BRAND_FONT",   "Poppins, Arial, Helvetica, sans-serif")


# =========================
# SCHEDULER SETTINGS
# =========================
# Default schedule (server local time)
SCHEDULE_TIMES_DEFAULT = ["09:15", "18:15"]

# How often to wake up and check time (seconds)
SCHEDULER_POLL_SECONDS = 20

# Prevent re-sending too frequently for the same scheduled slot
# If script runs continuously, it will send each slot once per day.
SEND_GUARD_MINUTES = 2


# =========================
# Helpers
# =========================

def _now_utc() -> datetime:
    return datetime.now(timezone.utc).replace(microsecond=0)


def _canon_status(material_status: str) -> str:
    # ✅ Exact FG checks pulled from app.py logic
    s0 = (material_status or "").strip().lower()
    s0 = " ".join(s0.split())
    
    fg_statuses = {
        "finished status",
        "tpi completed",
        "levelling completed",
        "offer to pfp/ssd",
        "quenching done",
    }
    
    if s0 in fg_statuses:
        return "FG"
    return "WIP"


def _is_dispatched(status: str) -> bool:
    return (status or "").strip().lower() == "dispatched"


def _pick_since_iso(added_at, created_at, updated_at=None) -> str:
    return (added_at or created_at or updated_at or "").strip()


def _parse_isoish(s: str) -> Optional[datetime]:
    if not s:
        return None
    raw = str(s).strip()
    try:
        # Date-only
        if len(raw) == 10 and raw[4] == "-" and raw[7] == "-":
            return datetime.fromisoformat(raw).replace(tzinfo=timezone.utc)

        # ISO with Z
        if raw.endswith("Z"):
            raw = raw[:-1] + "+00:00"

        dt = datetime.fromisoformat(raw)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except Exception:
        return None


def _age_days(since_iso: str, now_dt: datetime) -> Optional[int]:
    dt = _parse_isoish(since_iso)
    if not dt:
        return None
    diff = now_dt - dt
    if diff.total_seconds() < 0:
        return 0
    return int(diff.total_seconds() // 86400)


def _safe_float(v: Any, default: float = 0.0) -> float:
    try:
        if v is None:
            return default
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip()
        if not s or s.lower() == "null":
            return default
        return float(s)
    except Exception:
        return default


def _fmt_tons(x: float) -> str:
    # ✅ show 2 decimals and "MT"
    try:
        return f"{x:,.2f}"
    except Exception:
        return str(x)


@dataclass
class AlertRow:
    plate_id: str
    item_type: str
    bin: str
    status: str       # FG/WIP
    customer: str
    since_iso: str
    age_days: int
    level: str        # yellow/red
    weight: str


def _classify(status_norm: str, age_days: int) -> Optional[str]:
    if status_norm == "FG":
        if age_days > FG_RED_DAYS:
            return "red"
        if age_days > FG_YELLOW_DAYS:
            return "yellow"
        return None

    if status_norm == "WIP":
        if age_days > WIP_RED_DAYS:
            return "red"
        if age_days > WIP_YELLOW_DAYS:
            return "yellow"
        return None

    return None


# =========================
# DB URL resolution
# =========================

def resolve_db_url() -> str:
    db_url = (os.getenv("DATABASE_URL") or "").strip()
    if db_url:
        return db_url

    db_path = (os.getenv("YARD_DB_PATH") or "yard_logic/yard_data.db").strip()
    if not os.path.isabs(db_path):
        here = os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(here, db_path)

    return f"sqlite:///{db_path}"


# =========================
# DB Fetch
# =========================

def fetch_yard_items(engine) -> List[Dict[str, Any]]:
    # ✅ Included Material_Status in the DB fetch query
    sql = """
        SELECT
            plate_id,
            type,
            bin,
            status,
            Material_Status,
            customer,
            weight,
            added_at,
            created_at,
            updated_at
        FROM plates
        WHERE COALESCE(status,'') != 'Dispatched'
    """
    with engine.begin() as con:
        rows = con.execute(text(sql)).mappings().all()
    return [dict(r) for r in rows]


def build_alerts(rows: List[Dict[str, Any]]) -> Dict[str, List[AlertRow]]:
    now_dt = _now_utc()

    buckets: Dict[str, List[AlertRow]] = {
        "FG_red": [],
        "FG_yellow": [],
        "WIP_red": [],
        "WIP_yellow": [],
    }

    for r in rows:
        st_raw = r.get("status") or ""
        if _is_dispatched(st_raw):
            continue

        mat_status = r.get("Material_Status") or ""
        st_norm = _canon_status(mat_status)
        if st_norm not in ("FG", "WIP"):
            continue

        since_iso = _pick_since_iso(r.get("added_at"), r.get("created_at"), r.get("updated_at"))
        age = _age_days(since_iso, now_dt)
        if age is None:
            continue

        level = _classify(st_norm, age)
        if not level:
            continue

        ar = AlertRow(
            plate_id=str(r.get("plate_id") or "").strip(),
            item_type=str(r.get("type") or "").strip(),
            bin=str(r.get("bin") or "").strip().upper(),
            status=st_norm,
            customer=str(r.get("customer") or "").strip(),
            since_iso=since_iso,
            age_days=age,
            level=level,
            weight=str(r.get("weight") or "").strip(),
        )

        buckets[f"{st_norm}_{level}"].append(ar)

    for k in buckets:
        buckets[k].sort(key=lambda x: (-x.age_days, x.plate_id))

    return buckets


def _sum_bucket_weight(rows: List[AlertRow]) -> float:
    total = 0.0
    for r in rows:
        total += _safe_float(r.weight, 0.0)
    return total


# =========================
# WhatsApp summary (short)
# =========================

def build_whatsapp_text(b: Dict[str, List[AlertRow]]) -> str:
    fg_red = len(b["FG_red"])
    fg_yel = len(b["FG_yellow"])
    wip_red = len(b["WIP_red"])
    wip_yel = len(b["WIP_yellow"])

    total_red = fg_red + wip_red
    total_yel = fg_yel + wip_yel

    now_local = datetime.now().strftime("%Y-%m-%d %H:%M")

    # (WhatsApp kept same — no extra weight added here as per your request: only mail body cards)
    return (
        f"{ALERT_TITLE}\n"
        f"Generated: {now_local}\n\n"
        f"🔴 TOTAL RED: {total_red}\n"
        f"🟡 TOTAL YELLOW: {total_yel}\n\n"
        f"FG  -> 🔴 {fg_red} | 🟡 {fg_yel}\n"
        f"WIP -> 🔴 {wip_red} | 🟡 {wip_yel}\n\n"
        f"Rules:\n"
        f"FG >15d = Yellow, >20d = Red\n"
        f"WIP >3d = Yellow, >7d = Red\n"
    )


# =========================
# Excel attachment creation
# =========================

def _excel_styles():
    header_fill = PatternFill("solid", fgColor="0F2A43")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return header_fill, header_font, wrap, center, border


def _auto_fit_columns(ws, min_w=10, max_w=40):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        w = max(min_w, min(max_w, max_len + 2))
        ws.column_dimensions[col_letter].width = w


def build_excel_bytes(b: Dict[str, List[AlertRow]]) -> bytes:
    wb = Workbook()
    header_fill, header_font, wrap, center, border = _excel_styles()

    wb.remove(wb.active)

    def add_sheet(name: str, rows: List[AlertRow]):
        ws = wb.create_sheet(title=name)
        headers = ["ID", "Type", "Status", "Age(d)", "Bin", "Customer", "Weight", "Since"]
        ws.append(headers)

        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        for r in rows:
            ws.append([
                r.plate_id,
                r.item_type,
                r.status,
                r.age_days,
                r.bin,
                r.customer,
                r.weight,
                r.since_iso,
            ])

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = wrap
                cell.border = border

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:H{ws.max_row}"
        for rr in range(2, ws.max_row + 1):
            ws.row_dimensions[rr].height = 22

        _auto_fit_columns(ws, min_w=10, max_w=45)

    add_sheet("FG_RED", b["FG_red"])
    add_sheet("FG_YELLOW", b["FG_yellow"])
    add_sheet("WIP_RED", b["WIP_red"])
    add_sheet("WIP_YELLOW", b["WIP_yellow"])

    ws = wb.create_sheet(title="SUMMARY", index=0)

    fg_red = len(b["FG_red"])
    fg_yel = len(b["FG_yellow"])
    wip_red = len(b["WIP_red"])
    wip_yel = len(b["WIP_yellow"])
    total_red = fg_red + wip_red
    total_yel = fg_yel + wip_yel
    now_local = datetime.now().strftime("%Y-%m-%d %H:%M")

    ws.append(["Report", ALERT_TITLE])
    ws.append(["Generated", now_local])
    ws.append([])
    ws.append(["Metric", "Count"])
    ws.append(["TOTAL RED", total_red])
    ws.append(["TOTAL YELLOW", total_yel])
    ws.append([])
    ws.append(["FG RED", fg_red])
    ws.append(["FG YELLOW", fg_yel])
    ws.append(["WIP RED", wip_red])
    ws.append(["WIP YELLOW", wip_yel])
    ws.append([])
    ws.append(["Rules", ""])
    ws.append(["FG", "FG >15d = Yellow, >20d = Red"])
    ws.append(["WIP", "WIP >3d = Yellow, >7d = Red"])

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(1, ws.max_row + 1):
        for c in range(1, 3):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 55
    ws.freeze_panes = "A1"

    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# =========================
# Email Rendering (HTML)
# =========================

def build_email_html(b: Dict[str, List[AlertRow]]) -> str:
    fg_red = len(b["FG_red"])
    fg_yel = len(b["FG_yellow"])
    wip_red = len(b["WIP_red"])
    wip_yel = len(b["WIP_yellow"])
    total_red = fg_red + wip_red
    total_yel = fg_yel + wip_yel

    # ✅ weights for each bucket + totals
    fg_red_w = _sum_bucket_weight(b["FG_red"])
    fg_yel_w = _sum_bucket_weight(b["FG_yellow"])
    wip_red_w = _sum_bucket_weight(b["WIP_red"])
    wip_yel_w = _sum_bucket_weight(b["WIP_yellow"])

    total_red_w = fg_red_w + wip_red_w
    total_yel_w = fg_yel_w + wip_yel_w

    now_local = datetime.now().strftime("%Y-%m-%d %H:%M")

    def esc(s: str) -> str:
        return str(s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    gradient = f"linear-gradient(90deg, {BRAND_ORANGE}, {BRAND_GREEN}, {BRAND_BLUE})"

    return f"""
<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width">
  <title>{esc(ALERT_TITLE)}</title>
</head>
<body style="margin:0;padding:0;background:{BRAND_BG};font-family:{BRAND_FONT};color:{BRAND_TEXT};">
  <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:{BRAND_BG};padding:24px 12px;">
    <tr>
      <td align="center">
        <table role="presentation" width="820" cellspacing="0" cellpadding="0"
               style="max-width:820px;background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 10px 24px rgba(15,23,42,.10);">
          <tr>
            <td style="padding:18px 22px;">
              <div style="font-weight:800;font-size:18px;">{esc(ALERT_TITLE)}</div>
              <div style="margin-top:4px;font-size:12px;color:#64748b;">Generated: {esc(now_local)}</div>
              <div style="margin-top:6px;font-size:12px;color:#475569;font-weight:800;">
                Generated By Plate Mill Yard Management Software.
              </div>
              <div style="margin-top:4px;font-size:12px;color:#64748b;">
                Aging is calculated from the time the material is received in the yard.
              </div>
            </td>
          </tr>

          <tr><td style="height:4px;background:{BRAND_GREEN};background-image:{gradient};"></td></tr>

          <tr>
            <td style="padding:18px 22px;">
              <div style="display:flex;gap:12px;flex-wrap:wrap;">

                <div style="flex:2;min-width:320px;border:1px solid #e5e7eb;border-radius:10px;padding:12px;background:#f8fafc;">
                  <div style="font-weight:900;">Breakdown</div>
                  <div style="margin-top:8px;font-size:13px;line-height:1.6;">
                    <div>
                      <b>FG</b> →
                      🔴 {fg_red} <span style="color:#64748b;">({esc(_fmt_tons(fg_red_w))} MT)</span>
                      |
                      🟡 {fg_yel} <span style="color:#64748b;">({esc(_fmt_tons(fg_yel_w))} MT)</span>
                      <span style="color:#64748b;">(Yellow &gt;{FG_YELLOW_DAYS}d, Red &gt;{FG_RED_DAYS}d)</span>
                    </div>
                    <div>
                      <b>WIP</b> →
                      🔴 {wip_red} <span style="color:#64748b;">({esc(_fmt_tons(wip_red_w))} MT)</span>
                      |
                      🟡 {wip_yel} <span style="color:#64748b;">({esc(_fmt_tons(wip_yel_w))} MT)</span>
                      <span style="color:#64748b;">(Yellow &gt;{WIP_YELLOW_DAYS}d, Red &gt;{WIP_RED_DAYS}d)</span>
                    </div>
                  </div>
                </div>

                <div style="flex:1;min-width:240px;border:1px solid #e5e7eb;border-radius:10px;padding:12px;background:#fff;">
                  <div style="font-weight:900;color:#991b1b;">🔴 TOTAL RED</div>
                  <div style="font-size:28px;font-weight:900;margin-top:6px;">{total_red}</div>
                  <div style="margin-top:6px;font-size:12px;color:#64748b;font-weight:800;">Weight: {esc(_fmt_tons(total_red_w))} MT</div>
                </div>

                <div style="flex:1;min-width:240px;border:1px solid #e5e7eb;border-radius:10px;padding:12px;background:#fff;">
                  <div style="font-weight:900;color:#92400e;">🟡 TOTAL YELLOW</div>
                  <div style="font-size:28px;font-weight:900;margin-top:6px;">{total_yel}</div>
                  <div style="margin-top:6px;font-size:12px;color:#64748b;font-weight:800;">Weight: {esc(_fmt_tons(total_yel_w))} MT</div>
                </div>

              </div>

              <div style="margin-top:14px;font-size:12px;color:#475569;line-height:1.55;">
                <b>Rules</b><br>
                FG &gt;{FG_YELLOW_DAYS}d = Yellow, &gt;{FG_RED_DAYS}d = Red<br>
                WIP &gt;{WIP_YELLOW_DAYS}d = Yellow, &gt;{WIP_RED_DAYS}d = Red
              </div>

              <div style="margin-top:12px;font-size:12px;color:#6b7280;">
                Please find the attached Excel file for detailed item list (FG/WIP Red & Yellow) including weight.
              </div>

            </td>
          </tr>

          <tr>
            <td style="padding:14px 22px;border-top:1px solid #e5e7eb;font-size:12px;color:#6b7280;">
              © Jindal Steel. All rights reserved.
            </td>
          </tr>

        </table>
      </td>
    </tr>
  </table>
</body>
</html>
""".strip()


# =========================
# Send Email (HTML + Excel attachment)
# =========================

def send_email_html_with_attachment(
    subject: str,
    html_body: str,
    to_emails: List[str],
    cc_emails: List[str],
    attachment_bytes: bytes,
    attachment_filename: str,
) -> bool:
    if not to_emails:
        print("[email] No recipients. Skipping email send.")
        return False

    message = MIMEMultipart()
    # ✅ Set the custom display name to match exactly what you want
    message["From"] = f"{SENDER_NAME} <{SENDER_EMAIL}>"
    message["To"] = ", ".join(to_emails)
    if cc_emails:
        message["Cc"] = ", ".join(cc_emails)
    message["Subject"] = subject

    message.attach(MIMEText(html_body, "html"))

    part = MIMEApplication(
        attachment_bytes,
        _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    part.add_header("Content-Disposition", "attachment", filename=attachment_filename)
    message.attach(part)

    # ✅ IMPORTANT: send to both To + Cc
    all_rcpt = list(to_emails) + (list(cc_emails) if cc_emails else [])

    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT, timeout=30) as server:
            # We still authenticate and send as SENDER_EMAIL, but the recipient sees SENDER_NAME
            server.sendmail(SENDER_EMAIL, all_rcpt, message.as_string())
        print(f"[email] sent to {to_emails} cc {cc_emails} (with attachment: {attachment_filename})")
        return True
    except Exception as e:
        print("[email] failed:", repr(e))
        return False


# =========================
# WhatsApp (Twilio)
# =========================

def _format_whatsapp_to(india_number_or_plus: str) -> str:
    n = str(india_number_or_plus).strip()
    if n.startswith("+"):
        return "whatsapp:" + n
    return "whatsapp:+91" + n


def send_whatsapp(body: str, to_numbers: List[str]) -> bool:
    if TwilioClient is None:
        print("[wa] twilio not installed/importable. Skipping WhatsApp send.")
        return False

    sid = (os.getenv("TWILIO_ACCOUNT_SID") or "").strip()
    tok = (os.getenv("TWILIO_AUTH_TOKEN") or "").strip()
    wa_from = (os.getenv("TWILIO_WHATSAPP_FROM") or "").strip()

    if not (sid and tok and wa_from):
        print("[wa] Twilio WhatsApp not configured. Skipping WhatsApp send.")
        return False

    if not to_numbers:
        print("[wa] No recipients. Skipping WhatsApp send.")
        return False

    try:
        client = TwilioClient(sid, tok)
        ok_any = False
        for n in to_numbers:
            to = _format_whatsapp_to(n)
            client.messages.create(from_=wa_from, to=to, body=body)
            ok_any = True
            print(f"[wa] sent to {to}")
        return ok_any
    except Exception as e:
        print("[wa] failed:", repr(e))
        return False


# =========================
# Sending core
# =========================

def send_alert_once(engine, email_to: List[str], email_cc: List[str], whatsapp_to: List[str], dry_run: bool = False) -> Dict[str, Any]:
    rows = fetch_yard_items(engine)
    buckets = build_alerts(rows)

    fg_red = len(buckets["FG_red"])
    fg_yel = len(buckets["FG_yellow"])
    wip_red = len(buckets["WIP_red"])
    wip_yel = len(buckets["WIP_yellow"])

    subject = (
        f"{ALERT_TITLE} | "
        f"FG 🔴{fg_red} 🟡{fg_yel} | "
        f"WIP 🔴{wip_red} 🟡{wip_yel}"
    )

    wa_text = build_whatsapp_text(buckets)
    email_html = build_email_html(buckets)

    xlsx_bytes = build_excel_bytes(buckets)
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    xlsx_name = f"Yard_Aging_Details_{stamp}.xlsx"

    if dry_run:
        print(wa_text)
        return {
            "ok": True,
            "FG_yellow": fg_yel,
            "FG_red": fg_red,
            "WIP_yellow": wip_yel,
            "WIP_red": wip_red,
            "sent": False,
            "sent_email": False,
            "sent_whatsapp": False,
            "email_to": email_to,
            "email_cc": email_cc,
            "whatsapp_to": whatsapp_to,
            "attachment": xlsx_name
        }

    sent_email = send_email_html_with_attachment(
        subject=subject,
        html_body=email_html,
        to_emails=email_to,
        cc_emails=email_cc,
        attachment_bytes=xlsx_bytes,
        attachment_filename=xlsx_name,
    )
    sent_wa = send_whatsapp(body=wa_text, to_numbers=whatsapp_to)

    return {
        "ok": True,
        "FG_yellow": fg_yel,
        "FG_red": fg_red,
        "WIP_yellow": wip_yel,
        "WIP_red": wip_red,
        "sent": bool(sent_email or sent_wa),
        "sent_email": sent_email,
        "sent_whatsapp": sent_wa,
        "email_to": email_to,
        "email_cc": email_cc,
        "whatsapp_to": whatsapp_to,
        "attachment": xlsx_name
    }


# =========================
# Scheduler
# =========================

def _parse_times(times: List[str]) -> List[Tuple[int, int]]:
    out = []
    for t in times:
        s = (t or "").strip()
        if not s:
            continue
        hh, mm = s.split(":")
        out.append((int(hh), int(mm)))
    return out


def run_scheduler(engine, email_to: List[str], email_cc: List[str], whatsapp_to: List[str], times: List[str], dry_run: bool = False):
    """
    Keeps running and sends exactly once per scheduled slot per day.
    """
    slots = _parse_times(times)
    if not slots:
        print("[sched] No schedule times configured. Exiting.")
        return

    print(f"[sched] Running scheduler. Times={times} (local server time). Poll={SCHEDULER_POLL_SECONDS}s")

    # guard: { "YYYY-MM-DD_HH:MM": last_sent_epoch }
    last_sent: Dict[str, float] = {}

    while True:
        now = datetime.now()  # local server time
        day = now.strftime("%Y-%m-%d")
        hh = now.hour
        mm = now.minute

        for (H, M) in slots:
            if hh == H and mm == M:
                key = f"{day}_{H:02d}:{M:02d}"
                last_ts = last_sent.get(key, 0.0)
                age_minutes = (time.time() - last_ts) / 60.0 if last_ts else 999.0

                # ✅ Guard to avoid multiple sends during same minute / quick restarts
                if age_minutes >= SEND_GUARD_MINUTES:
                    print(f"[sched] Triggering send for slot {H:02d}:{M:02d} (key={key})")
                    try:
                        result = send_alert_once(engine, email_to, email_cc, whatsapp_to, dry_run=dry_run)
                        print(json.dumps(result))
                        last_sent[key] = time.time()
                    except Exception as e:
                        print("[sched] send failed:", repr(e))

        time.sleep(SCHEDULER_POLL_SECONDS)


# =========================
# Main
# =========================

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--run-once", action="store_true", help="Run once and exit (send immediately once)")
    ap.add_argument("--send-now", action="store_true", help="Send immediately once (same as --run-once)")
    ap.add_argument("--dry-run", action="store_true", help="Do not send, just print result")
    ap.add_argument("--daemon", action="store_true", help="Keep running and send at scheduled times")
    ap.add_argument(
        "--times",
        default=",".join(SCHEDULE_TIMES_DEFAULT),
        help="Comma separated schedule times in HH:MM (local server time). Default: 09:15,18:15"
    )
    args = ap.parse_args()

    db_url = resolve_db_url()
    engine = create_engine(db_url, future=True)

    wa_list = (os.getenv("ALERT_WHATSAPP_TO") or "").strip()
    email_list = (os.getenv("ALERT_EMAIL_TO") or "").strip()
    email_cc_list = (os.getenv("ALERT_EMAIL_CC") or "").strip()

    whatsapp_to = DEFAULT_WHATSAPP_NUMBERS if not wa_list else [x.strip() for x in wa_list.split(",") if x.strip()]
    email_to = DEFAULT_EMAIL_RECIPIENTS if not email_list else [x.strip() for x in email_list.split(",") if x.strip()]
    email_cc = DEFAULT_EMAIL_CC if not email_cc_list else [x.strip() for x in email_cc_list.split(",") if x.strip()]

    times = [t.strip() for t in (args.times or "").split(",") if t.strip()]

    # ✅ if user wants single immediate send:
    if args.run_once or args.send_now:
        result = send_alert_once(engine, email_to, email_cc, whatsapp_to, dry_run=args.dry_run)
        result["db"] = db_url
        print(json.dumps(result))
        return

    # ✅ default behavior: run scheduler if --daemon, else show help-like message
    if args.daemon:
        run_scheduler(engine, email_to, email_cc, whatsapp_to, times=times, dry_run=args.dry_run)
        return

    # If no args provided, we can default to daemon for convenience:
    # (You can change this behavior if you prefer strict usage.)
    print("[info] No mode selected. Starting scheduler by default.")
    run_scheduler(engine, email_to, email_cc, whatsapp_to, times=times, dry_run=args.dry_run)


if __name__ == "__main__":
    main()