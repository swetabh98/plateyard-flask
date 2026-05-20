from __future__ import annotations

import os
import csv
import io
import json
import time
import argparse
import urllib.request
from dataclasses import dataclass
from datetime import datetime, timezone
from math import floor
from typing import List, Dict, Any, Optional, Tuple

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

try:
    from twilio.rest import Client as TwilioClient
except Exception:
    TwilioClient = None


# =========================
# CONFIG
# =========================

ALERT_TITLE = "Plate Mill Yard Dispatchable Stock Report"

# WhatsApp kept same style as reference
DEFAULT_WHATSAPP_NUMBERS = ["9109951823"]  # India number without +91

# TO recipients
DEFAULT_EMAIL_RECIPIENTS = [
    "rupak.ranjan@jindalsteel.in",
    "pmlog.shiftincharge.rgh@jindalsteel.in",
    "hemant.patel1@jindalsteel.in",
    "kshitibhushan.gupta@jindalsteel.in",
    "sunilpatel.ex@gmail.com",
    "prashant.chaudhari@jindalsteel.in",
]

# CC recipients
DEFAULT_EMAIL_CC = [
    "lalit.goyal@jindalsteel.in",
    "swetabh.sinha@jindalsteel.in",
    "aritra.de@jindalsteel.in",
    "saurabh.agrawal@jindalsteel.in",
    "man.singh@jindalsteel.in",
    "sandeep.tyagi@jindalsteel.in",
    "vikram.singh1@jindalsteel.in",
    "sushil.mohanty@jindalsteel.in",
    "abhimanyu.singh@jindalsteel.in",
    "devendra.kalambe@jindalsteel.in",
]

# Google Sheet CSV URL
GOOGLE_SHEET_CSV_URL = os.environ.get(
    "YARD_SHEET_CSV_URL",
    "https://docs.google.com/spreadsheets/d/1TPt1wTmOFj4ydC_cGS59DGCf4enFGoe-9J3LW-nbRzE/export?format=csv",
)

# SMTP
SENDER_NAME = "Yard Management Software"
SENDER_EMAIL = "noreply.digital@jindalsteel.com"
SMTP_SERVER = "172.17.1.17"
SMTP_PORT = 25

# Branding
BRAND_ORANGE = os.environ.get("JINDAL_BRAND_ORANGE", "#F58220")
BRAND_GREEN = os.environ.get("JINDAL_BRAND_GREEN", "#3AAA35")
BRAND_BLUE = os.environ.get("JINDAL_BRAND_BLUE", "#0067B2")
BRAND_BG = os.environ.get("JINDAL_BRAND_MUTE", "#F4F6F8")
BRAND_TEXT = os.environ.get("JINDAL_BRAND_TEXT", "#333333")
BRAND_FONT = os.environ.get("JINDAL_BRAND_FONT", "Poppins, Arial, Helvetica, sans-serif")

# Scheduler
SCHEDULE_TIMES_DEFAULT = ["06:00", "14:00", "22:00"]
SCHEDULER_POLL_SECONDS = 20
SEND_GUARD_MINUTES = 2

# Allowed statuses
ALLOWED_MATERIAL_STATUSES = {
    "finished status",
    "tpi completed",
}


# =========================
# Helpers
# =========================

def _now_utc() -> datetime:
    return datetime.now(timezone.utc).replace(microsecond=0)


def _as_str(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _norm_text(v: Any) -> str:
    return " ".join(_as_str(v).split())


def _lower(v: Any) -> str:
    return _as_str(v).lower().strip()


def _safe_float(v: Any, default: float = 0.0) -> float:
    try:
        if v is None:
            return default
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip().replace(",", "")
        if not s or s.lower() == "null":
            return default
        return float(s)
    except Exception:
        return default


def _safe_int(v: Any, default: int = 0) -> int:
    try:
        if v is None:
            return default
        if isinstance(v, int):
            return v
        if isinstance(v, float):
            return int(v)
        s = str(v).strip().replace(",", "")
        if not s:
            return default
        return int(float(s))
    except Exception:
        return default


def _fmt_mt(x: float) -> str:
    try:
        return f"{x:,.2f}"
    except Exception:
        return str(x)


def _fetch_csv_rows(url: str, timeout: int = 45) -> List[Dict[str, str]]:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        raw = resp.read()

    txt = raw.decode("utf-8", errors="replace")
    head = txt[:200].lower()
    if "<html" in head and "google" in head:
        return []

    buf = io.StringIO(txt)
    return [{k: (v if v is not None else "") for k, v in r.items()} for r in csv.DictReader(buf)]


def _gget(row: Dict[str, str], key: str) -> Optional[str]:
    want = key.lower().replace(" ", "").replace("_", "").replace("-", "").replace(".", "")
    for k, v in row.items():
        kk = str(k).strip().lower().replace(" ", "").replace("_", "").replace("-", "").replace(".", "")
        if kk == want:
            return v
    return None


def _pick_any(row: Dict[str, str], keys: List[str]) -> str:
    for k in keys:
        v = _gget(row, k)
        if v is not None:
            return _as_str(v)
    return ""


def _extract_so_item(v: Any) -> str:
    s = _as_str(v)
    if "|" in s:
        parts = s.split("|")
        return _as_str(parts[-1])
    return s


@dataclass
class DispatchRow:
    material_status: str
    batch: str
    slocation: str
    bin_no: str
    v_thickness: str
    v_width: str
    v_length: str
    v_pieces: str
    qty: str
    v_ext_grade: str
    so_no: str
    so_item: str
    sold_to_party: str
    customer_name: str
    shipping_destination: str
    aging_days: str
    unres_stock: str
    bal2bill: str
    dispatchable_from_bal2bill: int
    party_trnsp_co_trnsp: str
    sold_to_party_2: str
    ship_to_party: str
    payment_status: str
    del_block: str
    planning_material: str
    qualityremark: str
    material: str
    sold_to_party_code: str


# =========================
# Google Sheet parsing
# =========================

def build_dispatchable_rows_from_sheet(csv_url: str) -> List[DispatchRow]:
    rows = _fetch_csv_rows(csv_url)
    out: List[DispatchRow] = []

    for r in rows:
        material_status = _pick_any(r, ["Material_Status", "Material Status"])
        if _lower(material_status) not in ALLOWED_MATERIAL_STATUSES:
            continue

        batch = _pick_any(r, ["Batch"])
        slocation = _pick_any(r, ["SLocation", "S Location"])
        bin_no = _pick_any(r, ["BinNo", "Batch Storage Bin", "BIN_NO", "BIN"])
        v_thickness = _pick_any(r, ["V_THICKNESS", "BATCH-V_THICKNESS"])
        v_width = _pick_any(r, ["V_WIDTH", "BATCH-V_WIDTH"])
        v_length = _pick_any(r, ["V_LENGTH", "BATCH-V_LENGTH"])
        v_pieces = _pick_any(r, ["V_PIECES", "BATCH-V_PIECES"])
        qty = _pick_any(r, ["Qty", "QTY", "Weight", "WEIGHT"])
        v_ext_grade = _pick_any(r, ["V_EXT_GRADE", "BATCH-P_EXT_GRADE", "P_EXT_GRADE"])
        so_no = _pick_any(r, ["SO No", "SO_NO", "SONo"])
        so_item_raw = _pick_any(r, ["SO_ITEM", "SO Item"])
        so_item = _extract_so_item(so_item_raw)
        sold_to_party = _pick_any(r, ["SoldToParty", "Sold-to Party"])
        customer_name = _pick_any(r, ["CustomerName", "Customer Name", "Customer"])
        shipping_destination = _pick_any(r, ["Shiping Destination", "Shipping Destination"])
        aging_days = _pick_any(r, ["Aging Days"])
        unres_stock = _pick_any(r, ["Unres. Stock", "Unres Stock", "Unres_Stock"])
        bal2bill = _pick_any(r, ["Bal2Bill", "Bal 2 Bill"])
        party_trnsp_co_trnsp = _pick_any(r, ["Party Trnsp/Co. Trnsp", "Party Trnsp Co Trnsp"])
        sold_to_party_2 = _pick_any(r, ["Sold to party"])
        ship_to_party = _pick_any(r, ["Ship to party"])
        payment_status = _pick_any(r, ["Payment Status"])
        del_block = _pick_any(r, ["FI_Rel_text", "FI Rel text", "Del. Block"])
        planning_material = _pick_any(r, ["Planning Material"])
        qualityremark = _pick_any(r, ["QUALITYREMARK", "BATCH-V_QUALITYREMARK", "V_QUALITYREMARK"])
        material = _pick_any(r, ["Material", "MaterialCode"])
        sold_to_party_code = _pick_any(r, ["Sold-to Party Code", "Sold to Party Code"])

        qty_f = _safe_float(qty, 0.0)
        pieces_i = _safe_int(v_pieces, 0)
        bal2bill_f = _safe_float(bal2bill, 0.0)

        if qty_f <= 0 or pieces_i <= 0 or bal2bill_f <= 0:
            continue

        single_piece_weight = qty_f / float(pieces_i)
        if single_piece_weight <= 0:
            continue

        dispatchable = int(floor(bal2bill_f / single_piece_weight))

        if dispatchable < 1:
            continue

        out.append(
            DispatchRow(
                material_status=material_status,
                batch=batch,
                slocation=slocation,
                bin_no=bin_no,
                v_thickness=v_thickness,
                v_width=v_width,
                v_length=v_length,
                v_pieces=v_pieces,
                qty=qty,
                v_ext_grade=v_ext_grade,
                so_no=so_no,
                so_item=so_item,
                sold_to_party=sold_to_party,
                customer_name=customer_name,
                shipping_destination=shipping_destination,
                aging_days=aging_days,
                unres_stock=unres_stock,
                bal2bill=bal2bill,
                dispatchable_from_bal2bill=dispatchable,
                party_trnsp_co_trnsp=party_trnsp_co_trnsp,
                sold_to_party_2=sold_to_party_2,
                ship_to_party=ship_to_party,
                payment_status=payment_status,
                del_block=del_block,
                planning_material=planning_material,
                qualityremark=qualityremark,
                material=material,
                sold_to_party_code=sold_to_party_code,
            )
        )

    out.sort(key=lambda x: (
        _lower(x.material_status),
        _lower(x.so_no),
        _lower(x.so_item),
        _lower(x.batch),
        _lower(x.material),
    ))
    return out


def _sum_qty(rows: List[DispatchRow]) -> float:
    total = 0.0
    for r in rows:
        total += _safe_float(r.qty, 0.0)
    return total


# =========================
# Excel attachment creation
# =========================

def _excel_styles():
    header_fill = PatternFill("solid", fgColor="0F2A43")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return header_fill, header_font, wrap, center, border


def _auto_fit_columns(ws, min_w=10, max_w=40):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        w = max(min_w, min(max_w, max_len + 2))
        ws.column_dimensions[col_letter].width = w


def build_excel_bytes(rows: List[DispatchRow]) -> bytes:
    wb = Workbook()
    header_fill, header_font, wrap, center, border = _excel_styles()

    ws = wb.active
    ws.title = "DISPATCHABLE_REPORT"

    headers = [
        "Material_Status",
        "Batch",
        "SLocation",
        "BinNo",
        "V_THICKNESS",
        "V_WIDTH",
        "V_LENGTH",
        "V_PIECES",
        "Qty",
        "V_EXT_GRADE",
        "SO No",
        "SO_ITEM",
        "SoldToParty",
        "CustomerName",
        "Shiping Destination",
        "Aging Days",
        "Unres. Stock",
        "Bal2Bill",
        "Dispatchable from Bal2Bill",
        "Party Trnsp/Co. Trnsp",
        "Sold to party",
        "Ship to party",
        "Payment Status",
        "Del. Block",
        "Planning Material",
        "QUALITYREMARK",
        "Material",
        "Sold-to Party Code",
    ]
    ws.append(headers)

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for r in rows:
        ws.append([
            r.material_status,
            r.batch,
            r.slocation,
            r.bin_no,
            r.v_thickness,
            r.v_width,
            r.v_length,
            r.v_pieces,
            r.qty,
            r.v_ext_grade,
            r.so_no,
            r.so_item,
            r.sold_to_party,
            r.customer_name,
            r.shipping_destination,
            r.aging_days,
            r.unres_stock,
            r.bal2bill,
            r.dispatchable_from_bal2bill,
            r.party_trnsp_co_trnsp,
            r.sold_to_party_2,
            r.ship_to_party,
            r.payment_status,
            r.del_block,
            r.planning_material,
            r.qualityremark,
            r.material,
            r.sold_to_party_code,
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = wrap
            cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    _auto_fit_columns(ws, min_w=12, max_w=28)

    summary = wb.create_sheet(title="SUMMARY", index=0)
    total_rows = len(rows)
    total_qty = _sum_qty(rows)
    total_dispatchable = sum(int(r.dispatchable_from_bal2bill) for r in rows)
    now_local = datetime.now().strftime("%Y-%m-%d %H:%M")

    summary.append(["Report", ALERT_TITLE])
    summary.append(["Generated", now_local])
    summary.append(["Source", "Google Sheet"])
    summary.append([])
    summary.append(["Metric", "Value"])
    summary.append(["Filtered statuses", "Finished Status, TPI completed"])
    summary.append(["Rows in report", total_rows])
    summary.append(["Total Qty", f"{_fmt_mt(total_qty)} MT"])
    summary.append(["Total Dispatchable from Bal2Bill", f"{total_dispatchable} Nos."])

    thin = Side(style="thin", color="CBD5E1")
    s_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(1, summary.max_row + 1):
        for c in range(1, 3):
            cell = summary.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = s_border

    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 60
    summary.freeze_panes = "A1"

    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# =========================
# Email Rendering (HTML)
# =========================

def build_email_html(rows: List[DispatchRow]) -> str:
    total_rows = len(rows)
    total_qty = _sum_qty(rows)
    total_dispatchable = sum(int(r.dispatchable_from_bal2bill) for r in rows)
    now_local = datetime.now().strftime("%Y-%m-%d %H:%M")

    def esc(s: str) -> str:
        return str(s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    gradient = f"linear-gradient(90deg, {BRAND_ORANGE}, {BRAND_GREEN}, {BRAND_BLUE})"

    return f"""
<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width">
  <title>{esc(ALERT_TITLE)}</title>
</head>
<body style="margin:0;padding:0;background:{BRAND_BG};font-family:{BRAND_FONT};color:{BRAND_TEXT};">
  <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:{BRAND_BG};padding:24px 12px;">
    <tr>
      <td align="center">
        <table role="presentation" width="820" cellspacing="0" cellpadding="0"
               style="max-width:820px;background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 10px 24px rgba(15,23,42,.10);">
          <tr>
            <td style="padding:18px 22px;">
              <div style="font-weight:800;font-size:18px;">{esc(ALERT_TITLE)}</div>
              <div style="margin-top:4px;font-size:12px;color:#64748b;">Generated: {esc(now_local)}</div>
              <div style="margin-top:6px;font-size:12px;color:#475569;font-weight:800;">
                Report Generated from Plate Mill Yard Management Software.
              </div>
            </td>
          </tr>

          <tr><td style="height:4px;background:{BRAND_GREEN};background-image:{gradient};"></td></tr>

          <tr>
            <td style="padding:18px 22px;">
              <div style="display:flex;gap:12px;flex-wrap:wrap;">

                <div style="flex:2;min-width:320px;border:1px solid #e5e7eb;border-radius:10px;padding:12px;background:#f8fafc;">
                  <div style="font-weight:900;">Report Summary</div>
                  <div style="margin-top:8px;font-size:13px;line-height:1.7;">
                    <div><b>Statuses included:</b> Finished Status, TPI completed</div>
                    <div><b>Rows in report:</b> {total_rows}</div>
                    <div><b>Total Qty:</b> {esc(_fmt_mt(total_qty))} MT</div>
                    <div><b>Total Dispatchable from Bal2Bill:</b> {total_dispatchable} Nos.</div>
                  </div>
                </div>

                <div style="flex:1;min-width:240px;border:1px solid #e5e7eb;border-radius:10px;padding:12px;background:#fff;">
                  <div style="font-weight:900;color:#0f2a43;">Rows Included</div>
                  <div style="font-size:28px;font-weight:900;margin-top:6px;">{total_rows}</div>
                </div>

                <div style="flex:1;min-width:240px;border:1px solid #e5e7eb;border-radius:10px;padding:12px;background:#fff;">
                  <div style="font-weight:900;color:#0f2a43;">Dispatchable Qty</div>
                  <div style="font-size:28px;font-weight:900;margin-top:6px;">{total_dispatchable} Nos.</div>
                </div>

              </div>

              <div style="margin-top:12px;font-size:12px;color:#6b7280;">
                Please find the attached Excel file for detailed dispatchable stock prepared from the Google Sheet.
              </div>

            </td>
          </tr>

          <tr>
            <td style="padding:14px 22px;border-top:1px solid #e5e7eb;font-size:12px;color:#6b7280;">
              © Jindal Steel. All rights reserved.
            </td>
          </tr>

        </table>
      </td>
    </tr>
  </table>
</body>
</html>
""".strip()


# =========================
# Send Email
# =========================

def send_email_html_with_attachment(
    subject: str,
    html_body: str,
    to_emails: List[str],
    cc_emails: List[str],
    attachment_bytes: bytes,
    attachment_filename: str,
) -> bool:
    if not to_emails:
        print("[email] No recipients. Skipping email send.")
        return False

    message = MIMEMultipart()
    # ✅ Set the custom display name to match exactly what you want
    message["From"] = f"{SENDER_NAME} <{SENDER_EMAIL}>"
    message["To"] = ", ".join(to_emails)
    if cc_emails:
        message["Cc"] = ", ".join(cc_emails)
    message["Subject"] = subject

    message.attach(MIMEText(html_body, "html"))

    part = MIMEApplication(
        attachment_bytes,
        _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    part.add_header("Content-Disposition", "attachment", filename=attachment_filename)
    message.attach(part)

    all_rcpt = list(to_emails) + (list(cc_emails) if cc_emails else [])

    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT, timeout=30) as server:
            server.sendmail(SENDER_EMAIL, all_rcpt, message.as_string())
        print(f"[email] sent to {to_emails} cc {cc_emails} (with attachment: {attachment_filename})")
        return True
    except Exception as e:
        print("[email] failed:", repr(e))
        return False


# =========================
# WhatsApp
# =========================

def build_whatsapp_text(rows: List[DispatchRow]) -> str:
    now_local = datetime.now().strftime("%Y-%m-%d %H:%M")
    total_rows = len(rows)
    total_qty = _sum_qty(rows)
    total_dispatchable = sum(int(r.dispatchable_from_bal2bill) for r in rows)

    return (
        f"{ALERT_TITLE}\n"
        f"Generated: {now_local}\n\n"
        f"Rows in report: {total_rows}\n"
        f"Total Qty: {_fmt_mt(total_qty)} MT\n"
        f"Dispatchable from Bal2Bill: {total_dispatchable} Nos.\n\n"
        f"Statuses: Finished Status, TPI completed\n"
    )


def _format_whatsapp_to(india_number_or_plus: str) -> str:
    n = str(india_number_or_plus).strip()
    if n.startswith("+"):
        return "whatsapp:" + n
    return "whatsapp:+91" + n


def send_whatsapp(body: str, to_numbers: List[str]) -> bool:
    if TwilioClient is None:
        print("[wa] twilio not installed/importable. Skipping WhatsApp send.")
        return False

    sid = (os.getenv("TWILIO_ACCOUNT_SID") or "").strip()
    tok = (os.getenv("TWILIO_AUTH_TOKEN") or "").strip()
    wa_from = (os.getenv("TWILIO_WHATSAPP_FROM") or "").strip()

    if not (sid and tok and wa_from):
        print("[wa] Twilio WhatsApp not configured. Skipping WhatsApp send.")
        return False

    if not to_numbers:
        print("[wa] No recipients. Skipping WhatsApp send.")
        return False

    try:
        client = TwilioClient(sid, tok)
        ok_any = False
        for n in to_numbers:
            to = _format_whatsapp_to(n)
            client.messages.create(from_=wa_from, to=to, body=body)
            ok_any = True
            print(f"[wa] sent to {to}")
        return ok_any
    except Exception as e:
        print("[wa] failed:", repr(e))
        return False


# =========================
# Sending core
# =========================

def send_alert_once(email_to: List[str], email_cc: List[str], whatsapp_to: List[str], dry_run: bool = False) -> Dict[str, Any]:
    rows = build_dispatchable_rows_from_sheet(GOOGLE_SHEET_CSV_URL)

    total_rows = len(rows)
    total_qty = _sum_qty(rows)
    total_dispatchable = sum(int(r.dispatchable_from_bal2bill) for r in rows)
    now_local = datetime.now().strftime("%Y-%m-%d %H:%M")

    subject = f"Dispatchable Stock Report | {now_local}"

    wa_text = build_whatsapp_text(rows)
    email_html = build_email_html(rows)

    xlsx_bytes = build_excel_bytes(rows)
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    xlsx_name = f"Dispatchable_Stock_Report_{stamp}.xlsx"

    if dry_run:
        with open(xlsx_name, "wb") as f:
            f.write(xlsx_bytes)
        print(f"[dry-run] Excel saved to: {os.path.abspath(xlsx_name)}")
        print(wa_text)
        return {
            "ok": True,
            "rows": total_rows,
            "total_qty": total_qty,
            "dispatchable": total_dispatchable,
            "sent": False,
            "sent_email": False,
            "sent_whatsapp": False,
            "email_to": email_to,
            "email_cc": email_cc,
            "whatsapp_to": whatsapp_to,
            "attachment": xlsx_name,
        }

    sent_email = send_email_html_with_attachment(
        subject=subject,
        html_body=email_html,
        to_emails=email_to,
        cc_emails=email_cc,
        attachment_bytes=xlsx_bytes,
        attachment_filename=xlsx_name,
    )
    sent_wa = send_whatsapp(body=wa_text, to_numbers=whatsapp_to)

    return {
        "ok": True,
        "rows": total_rows,
        "total_qty": total_qty,
        "dispatchable": total_dispatchable,
        "sent": bool(sent_email or sent_wa),
        "sent_email": sent_email,
        "sent_whatsapp": sent_wa,
        "email_to": email_to,
        "email_cc": email_cc,
        "whatsapp_to": whatsapp_to,
        "attachment": xlsx_name,
    }


# =========================
# Scheduler
# =========================

def _parse_times(times: List[str]) -> List[Tuple[int, int]]:
    out = []
    for t in times:
        s = (t or "").strip()
        if not s:
            continue
        hh, mm = s.split(":")
        out.append((int(hh), int(mm)))
    return out


def run_scheduler(email_to: List[str], email_cc: List[str], whatsapp_to: List[str], times: List[str], dry_run: bool = False):
    slots = _parse_times(times)
    if not slots:
        print("[sched] No schedule times configured. Exiting.")
        return

    print(f"[sched] Running scheduler. Times={times} (local server time). Poll={SCHEDULER_POLL_SECONDS}s")

    last_sent: Dict[str, float] = {}

    while True:
        now = datetime.now()
        day = now.strftime("%Y-%m-%d")
        hh = now.hour
        mm = now.minute

        for (H, M) in slots:
            if hh == H and mm == M:
                key = f"{day}_{H:02d}:{M:02d}"
                last_ts = last_sent.get(key, 0.0)
                age_minutes = (time.time() - last_ts) / 60.0 if last_ts else 999.0

                if age_minutes >= SEND_GUARD_MINUTES:
                    print(f"[sched] Triggering send for slot {H:02d}:{M:02d} (key={key})")
                    try:
                        result = send_alert_once(email_to, email_cc, whatsapp_to, dry_run=dry_run)
                        print(json.dumps(result))
                        last_sent[key] = time.time()
                    except Exception as e:
                        print("[sched] send failed:", repr(e))

        time.sleep(SCHEDULER_POLL_SECONDS)


# =========================
# Main
# =========================

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--run-once", action="store_true", help="Run once and exit (send immediately once)")
    ap.add_argument("--send-now", action="store_true", help="Send immediately once (same as --run-once)")
    ap.add_argument("--dry-run", action="store_true", help="Do not send, just print result")
    ap.add_argument("--daemon", action="store_true", help="Keep running and send at scheduled times")
    ap.add_argument(
        "--times",
        default=",".join(SCHEDULE_TIMES_DEFAULT),
        help="Comma separated schedule times in HH:MM (local server time). Default: 06:00,14:00,22:00"
    )
    args = ap.parse_args()

    wa_list = (os.getenv("ALERT_WHATSAPP_TO") or "").strip()
    email_list = (os.getenv("ALERT_EMAIL_TO") or "").strip()
    email_cc_list = (os.getenv("ALERT_EMAIL_CC") or "").strip()

    whatsapp_to = DEFAULT_WHATSAPP_NUMBERS if not wa_list else [x.strip() for x in wa_list.split(",") if x.strip()]
    email_to = DEFAULT_EMAIL_RECIPIENTS if not email_list else [x.strip() for x in email_list.split(",") if x.strip()]
    email_cc = DEFAULT_EMAIL_CC if not email_cc_list else [x.strip() for x in email_cc_list.split(",") if x.strip()]

    times = [t.strip() for t in (args.times or "").split(",") if t.strip()]

    if args.run_once or args.send_now:
        result = send_alert_once(email_to, email_cc, whatsapp_to, dry_run=args.dry_run)
        print(json.dumps(result))
        return

    if args.daemon:
        run_scheduler(email_to, email_cc, whatsapp_to, times=times, dry_run=args.dry_run)
        return

    print("[info] No mode selected. Starting scheduler by default.")
    run_scheduler(email_to, email_cc, whatsapp_to, times=times, dry_run=args.dry_run)


if __name__ == "__main__":
    main()